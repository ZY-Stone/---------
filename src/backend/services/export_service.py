"""
backend/services/export_service.py — 数据导出服务
"""
import io
import json
from datetime import datetime
from sqlalchemy.orm import Session
from models.sales_data import SalesWidth, SalesPotential
from models.import_record import ImportRecord
from models.audit_log import AuditLog
from models.product_dict import ProductDict


def export_width_excel(db: Session, user_info: dict) -> bytes:
    """导出产品宽度数据为 Excel (.xlsx)"""
    try:
        from openpyxl import Workbook
    except ImportError:
        return _export_csv_fallback(db, user_info, "width")

    wb = Workbook()
    ws = wb.active
    ws.title = "产品宽度数据"
    headers = ["期间", "客户名称", "最终用户", "产品", "销售额(万)", "同期销售额(万)", "数量", "同期数量", "是否规上"]
    ws.append(headers)

    query = db.query(SalesWidth).filter(SalesWidth.tenant_id == user_info.get("tenant_id", 1))
    for row in query.limit(10000).all():
        prod = db.query(ProductDict).filter(ProductDict.id == row.product_id).first()
        ws.append([
            row.period, row.customer_name, row.user_name or "",
            prod.name if prod else str(row.product_id),
            row.amount, row.amount_prev, row.qty, row.qty_prev,
            "是" if row.is_regulated else "否"
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_potential_excel(db: Session, user_info: dict) -> bytes:
    """导出潜力产品数据为 Excel"""
    try:
        from openpyxl import Workbook
    except ImportError:
        return _export_csv_fallback(db, user_info, "potential")

    wb = Workbook()
    ws = wb.active
    ws.title = "潜力产品数据"
    headers = ["期间", "客户名称", "最终用户", "产品", "销售额(万)", "同期销售额(万)", "数量", "同期数量", "商机数", "同期商机数"]
    ws.append(headers)

    query = db.query(SalesPotential).filter(SalesPotential.tenant_id == user_info.get("tenant_id", 1))
    for row in query.limit(10000).all():
        prod = db.query(ProductDict).filter(ProductDict.id == row.product_id).first()
        ws.append([
            row.period, row.customer_name, row.user_name or "",
            prod.name if prod else str(row.product_id),
            row.amount, row.amount_prev, row.qty, row.qty_prev,
            row.opps, row.opps_prev,
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_audit_excel(db: Session, user_info: dict) -> bytes:
    """导出审计日志为 Excel"""
    try:
        from openpyxl import Workbook
    except ImportError:
        return b""

    wb = Workbook()
    ws = wb.active
    ws.title = "审计日志"
    ws.append(["时间", "操作人ID", "操作类型", "操作对象", "详情", "IP"])

    logs = db.query(AuditLog).filter(AuditLog.tenant_id == user_info.get("tenant_id", 1)).order_by(
        AuditLog.created_at.desc()).limit(5000).all()
    for log in logs:
        ws.append([
            str(log.created_at) if log.created_at else "",
            log.user_id, log.action, log.target or "", log.detail or "", log.ip or ""
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _export_csv_fallback(db: Session, user_info: dict, dtype: str) -> bytes:
    """纯 CSV 兜底导出（不需要 openpyxl）"""
    lines = []
    if dtype == "width":
        lines.append("期间,客户名称,最终用户,产品ID,销售额(万),同期销售额(万),数量,同期数量,是否规上")
        query = db.query(SalesWidth).filter(SalesWidth.tenant_id == user_info.get("tenant_id", 1)).limit(10000)
        for row in query.all():
            lines.append(f"{row.period},{row.customer_name},{row.user_name or ''},{row.product_id},{row.amount},{row.amount_prev},{row.qty},{row.qty_prev},{'是' if row.is_regulated else '否'}")
    else:
        lines.append("期间,客户名称,最终用户,产品ID,销售额(万),同期销售额(万),数量,同期数量,商机数,同期商机数")
        query = db.query(SalesPotential).filter(SalesPotential.tenant_id == user_info.get("tenant_id", 1)).limit(10000)
        for row in query.all():
            lines.append(f"{row.period},{row.customer_name},{row.user_name or ''},{row.product_id},{row.amount},{row.amount_prev},{row.qty},{row.qty_prev},{row.opps},{row.opps_prev}")
    return "\n".join(lines).encode("utf-8-sig")
